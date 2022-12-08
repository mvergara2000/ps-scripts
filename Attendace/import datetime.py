from datetime import datetime
from datetime import timedelta
import os
import PyPDF2
import io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import( 
    PieChart,
    Reference
)
import pandas as pd
import math
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import time
from selenium.webdriver.chrome.options import Options
import os
import pandas as pd
from datetime import datetime
from datetime import timedelta
import json
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from cryptography.fernet import Fernet
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import PyPDF2
import warnings
import io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart3D, Reference
from openpyxl.utils.exceptions import WorkbookAlreadySaved

import math

def processPDF():
    
    today=datetime.now()
    todayFormatted=today.strftime('%m/%d/%Y')
    forFile = today.strftime(('%m%d%Y'))
    df=pd.DataFrame(columns = ['Date', 'Grade', 'Classroom', 'Day Membership Possible','# of students Absent', 'Day Attendance','# of students Attending'])
    if os.path.exists('classRoomAttendanceIncentive.xlsx'):
        print('The File Already Exists')
        num=33
    else:
        file_name = "classRoomAttendanceIncentive.xlsx"
        df.to_excel(file_name)
        num=1
    hn=os.environ["USERNAME"]
    path1= "C:/Users/"+hn+ "/Downloads/"
    #writer = pd.ExcelWriter('classRoomAttendanceIncentive.xlsx', engine='openpyxl')
    #wb = load_workbook(filename='classRoomAttendanceIncentive.xlsx')
    


    os.chdir(path1)#change path to downloads folder
#read pdf file, change if download file is different
    pdfFile= open("ClassAttendanceAudit.pdf", 'rb') 
    pdfReader= PyPDF2.PdfFileReader(pdfFile)
    pages = pdfReader.numPages
    t=0
    membership=0
    attendance=0
    absent=0

    for x in range(pages):
        pageObj = pdfReader.getPage(x)
        lines = pageObj.extractText().splitlines()
    
        for line in lines:
       
            #print(todayFormatted)
        
        
            if line.split()[0]=="Teacher:":#check line is teacher, get name
                if t == 0:
               
                    teacher = line.split()[1].replace(',','')
                
                    t+=1 
                

            if line.split()[0]=="Section:":#if line has section, get section name
           
                if t == 1:
                    grade=line.split()[1]

                    t+=1
                    
            if line.split()[0]=="Total":#if line has total get total membership and total attendance
                #print(line.split()[2])
                if line.split()[1]== "Membership:":
                    membership=int(line.split()[2])
                elif line.split()[1]== "Attendance:":
                    attendance=int(line.split()[2])
                    #print(y)
                    absent = membership-attendance 
                    #"# of students Absent: "+ 
                    percent= attendance/membership
                    percent=percent*100
                    if percent>90:
                        abovex=1
                    else:
                        abovex=0
                    df = df.append({'Date':todayFormatted, 'Classroom':teacher,'Grade':grade,'Day Membership Possible':membership,'# of students Absent': absent, 'Day Attendance':attendance,'# of students Attending':math.ceil(percent), 'Above 90': abovex},ignore_index = True)
                #print(str(absent)) , "% "+ "days above "+ login['threshold']: per
                t=0
    pdfFile.close()
    os.remove("ClassAttendanceAudit.pdf")
    os.chdir(os.path.realpath(os.path.join(os.path.dirname(__file__))))
    
    teacherList=df['Classroom'].values.tolist()
    gradeList=df['Grade'].values.tolist()
    above=df['Above 90'].values.tolist()
    membership= df['Day Membership Possible'].values
    if os.path.exists("classRoomAttendanceIncentive.xlsx"):
        print("file does exist")
    reader = pd.read_excel("classRoomAttendanceIncentive.xlsx", engine="openpyxl")
    book=load_workbook("classRoomAttendanceIncentive.xlsx")
    x=1
    if num == 33:
        book.create_sheet('forFile')
        worksheet=book['forFile']
        for names in teacherList:
            cell = 'A'+ str(x)
            cellb = 'B'+str(x)
            cellc= 'C'+str(x)
            celld= 'D'+ str(x)
            cellc= 'C'+str(x)
            celle= 'E'+ str(x)
            cellf= 'F'+str(x)
            cellg= 'G'+str(x)
            cellh= 'H'+str(x)
            
            worksheet[cell]= todayFormatted
            worksheet[cellb]= gradeList[x-1]
            worksheet[cellc]=names
            worksheet[celld]=df['Day Membership Possible'].values[x-1]
            worksheet[celle]=df['# of students Absent'].values[x-1]
            worksheet[cellf]=df['Day Attendance'].values[x-1]
            worksheet[cellg]=df['# of students Attending'].values[x-1]
            worksheet[cellh]= above[x-1]
            x+=1
            #
        worksheet['A1']= 'Date'
        worksheet['B1']= 'Grade'
        worksheet['C1']= 'Classroom'
        worksheet['D1']= 'Day Membership Possible'
        worksheet['E1']= '# of students Absent'
        worksheet['F1']= 'Day Attendance'
        worksheet['G1']= '# of students Attending'
        worksheet['H1']= 'Above 90'
          
    else:
        if not forFile in book.sheetnames:
            book.create_sheet(forFile)
        #book.create_sheet('Graph By Teacher')
        try:
            book.remove_sheet(book['Sheet1'])   
        except:
             pass     
        with pd.ExcelWriter("classRoomAttendanceIncentive.xlsx", engine="openpyxl") as writer:
            writer.book = book
            writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))
        
        
            
            df.to_excel(writer,sheet_name = forFile,index=False, startrow=len(reader)+num)
    #if not 'Graph by Teacher' in book.sheetnames:
    try:       
        
        worksheet = book['Graph By Teacher'] 
    except KeyError:
        print('Creating Worksheet')
        book.create_sheet('Graph By Teacher')
        book.save('classRoomAttendanceIncentive.xlsx')
        worksheet = book['Graph By Teacher'] 
        x=2
        for names in teacherList:
            cell = 'A'+ str(x)
            cellb = 'B'+str(x)
            cellc= 'C'+str(x)
            celle= 'E'+ str(x)
            worksheet[cell]=names
            worksheet[cellb]=0
            worksheet[cellc]=0
            worksheet[celle]=gradeList[x-2]
            x+=1
            #
        worksheet['B1']= 'Number of days above 90'
        worksheet['A1']= 'Class Room'
        worksheet['C1']= 'Total school Days'
        worksheet['D1']= '% days above 90'
        worksheet['E1']= 'Grade'


        book.save('classRoomAttendanceIncentive.xlsx')

    x=2
    for days in above:
            
        cellb = 'B'+str(x)
        cellc= 'C'+str(x)
        celld= 'D'+str(x)
        cellBval=worksheet[cellb].value
        cellCval=worksheet[cellc].value
        worksheet[cellb]=cellBval+above[x-2]
        worksheet[cellc]= cellCval+1
        worksheet[celld]=  (worksheet[cellb].value/ worksheet[cellc].value)*100

        x+=1
        book.save('classRoomAttendanceIncentive.xlsx')
    
    values= Reference(worksheet,
                      min_col=4,
                      max_col=4,
                      min_row=2,
                      max_row=34)
    cats = Reference(worksheet, min_col=1,max_col=1, min_row=2,max_row=34)
    chart = BarChart3D()
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    chart.title="Total Classroom Attendance %"
    chart.x_axis.title="Classroom"
    chart.x_axis.title="% Days Above 90"
    worksheet.add_chart(chart,"G1")
    book.save('classRoomAttendanceIncentive.xlsx')
    book.close()
  

if __name__=="__main__":
    processPDF()
    
