from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import time
from selenium.webdriver.chrome.options import Options
import os
import pandas as pd
from datetime import datetime
from datetime import timedelta
import json
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from cryptography.fernet import Fernet
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import PyPDF2
import warnings
import io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart3D, Reference
from openpyxl.utils.exceptions import WorkbookAlreadySaved

import math

def main():
    warnings.filterwarnings("ignore")
    chrome_options= webdriver.ChromeOptions()
    chrome_options.add_argument("--headless=chrome")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option('prefs',  {
    
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True
    }
)
 #generate key and open key.key

#opens key.key and assigns the key stored as key

# 
 
 #chrome_prefs = {"download.default_directory": r"C:\path\to\Downloads"} # (windows)
 #chrome_options.experimental_options["prefs"] = chrome_prefs

    
    ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__)))
    
    today=datetime.now()
    todayFormatted=today.strftime('%m/%d/%Y')
    #print(todayFormatted)
    if os.path.exists('filekey.key'):
        k = open('filekey.key')
        key = k.read()
        

    if os.path.exists('pslogin.json'):
        print("Logging In...")
    else:
        key = getLogin()
        encrypt(key)
    decrypt(key)
    f = open('pslogin.json')
    login= json.load(f)
    
    driver = webdriver.Chrome(options=chrome_options)
 #
 # url = "https://powerschool.npd117.net/admin/pw.html"
 
    logIn(driver,login)
    
    
    encrypt(key)
    hn=os.environ["USERNAME"]
    path1= "C:/Users/"+hn+ "/Downloads/"
    findSchool(driver,login)
    getabsentees(driver,todayFormatted)
  
    #export(driver)
        
    
 #print(grade)
    time.sleep(10)
    processPDF(login)
    f.close()
    #convertToXlsx(path1, grade, filePos,ROOT_DIR)


def getLogin():
   
    username=input('Enter Username: ')
    password=input('Enter Password: ')
    school = input('Enter school(GO or OR): ')
    threshold = input('Enter Percentage to stay above: ')
    if school == 'OR' or 'or':
        school= 'school_5'
    elif school == 'GO' or 'go':
        school='school_4'
    elif school == 'CN' or 'cn':
        school='school_3'
    dictionary = {
        "username": username,
        "password": password,
        
        "school": school,
        "threshold": threshold
        
    }
    
    json_object = json.dumps(dictionary, indent=4)
    with open("pslogin.json", "w") as outfile:
     outfile.write(json_object) 
    
    key = Fernet.generate_key()
    
 
# string the key in a file
    with open('filekey.key', 'wb') as filekey:
        filekey.write(key)
    return key
    
 
        
def encrypt(key):
    # opening the key
    with open('filekey.key', 'rb') as filekey:
        key = filekey.read()
 
    # using the generated key
    fernet = Fernet(key)
 
    # opening the original file to encrypt
    with open('pslogin.json', 'rb') as file:
        original = file.read()
     
    # encrypting the file
    encrypted = fernet.encrypt(original)
 
# opening the file in write mode and
# writing the encrypted data
    with open('pslogin.json', 'wb') as encrypted_file:
        encrypted_file.write(encrypted)

def decrypt(key):
    # using the key
    fernet = Fernet(key)
 # opening the encrypted file
    with open('pslogin.json', 'rb') as enc_file:
        encrypted = enc_file.read()
 # decrypting the file
    decrypted = fernet.decrypt(encrypted)
# opening the file in write mode and
# writing the decrypted data
    with open('pslogin.json', 'wb') as dec_file:
        dec_file.write(decrypted)
def logIn(driver, login):
    #uses user input to login
    
   
    userName=login['username']
    passWord = login['password']
    
    driver.get('https://powerschool.npd117.net/admin/pw.html')
    input= driver.find_element("id","fieldUsername")
    input.send_keys(userName)
    input= driver.find_element("id","fieldPassword")
    input.send_keys(passWord)
    input.submit()
    time.sleep(3)#change this to wait for element to load

def getabsentees(driver,date):
    url = "https://powerschool.npd117.net/admin/reports_engine/report_w_param.html?ac=reports_get_using_ID;repo_ID=PSPRE_CLASS_AUDIT"
    driver.get(url)
    
    radio=driver.find_elements(By.NAME,"param_reportDates")
    radio[1].click()
    
    day=driver.find_element(By.NAME,"param_startdate")
    day.clear()
    day.send_keys(date)#start date
    day=driver.find_element(By.NAME,"param_enddate")
    day.clear()
    day.send_keys(date)#start date
    
   
  
    driver.find_element("name","param_cb9;1").click()
    driver.find_element("id","btnSubmit").click()
    url = driver.current_url
    time.sleep(20)
    driver.find_element("id","prReloadButton").click()
    driver.find_element(By.LINK_TEXT,"View").click()
    save=ActionChains(driver)
    save.key_down(Keys.CONTROL).send_keys('S').key_up(Keys.CONTROL).perform()
    time.sleep(10)
    time.sleep(10)
 
def findSchool(driver,login):
    school = login['school']
    print("Finding school...")
    url = driver.current_url
    driver.get(url)
    schoolpicker=driver.find_element("id","adminSchoolPicker")
    schoolpicker.click()
    time.sleep(4)
    schoolpicker=driver.find_element("id",school)
    schoolpicker.click()
    time.sleep(4) 


def processPDF(login):
    today=datetime.now()
    todayFormatted=today.strftime('%m/%d/%Y')
    forFile = today.strftime(('%m%d%Y'))
    df=pd.DataFrame(columns = ['Date', 'Grade', 'Classroom', 'Day Membership Possible','# of students Absent', 'Day Attendance','# of students Attending'])
    if os.path.exists('classRoomAttendanceIncentive.xlsx'):
        print('The File Already Exists')
        num=33
    else:
        file_name = "classRoomAttendanceIncentive.xlsx"
        df.to_excel(file_name)
        num=1
    hn=os.environ["USERNAME"]
    path1= "C:/Users/"+hn+ "/Downloads/"
    #writer = pd.ExcelWriter('classRoomAttendanceIncentive.xlsx', engine='openpyxl')
    #wb = load_workbook(filename='classRoomAttendanceIncentive.xlsx')
    reader = pd.read_excel("classRoomAttendanceIncentive.xlsx", engine="openpyxl")
    book=load_workbook("classRoomAttendanceIncentive.xlsx")


    os.chdir(path1)
    pdfFile= open("ClassAttendanceAudit.pdf", 'rb') 
    pdfReader= PyPDF2.PdfFileReader(pdfFile)
    pages = pdfReader.numPages
    t=0
    membership=0
    attendance=0
    absent=0

    for x in range(pages):
        pageObj = pdfReader.getPage(x)
        lines = pageObj.extractText().splitlines()
    
        for line in lines:
       
            #print(todayFormatted)
        
        
            if line.split()[0]=="Teacher:":
                if t == 0:
               
                    teacher = line.split()[1].replace(',','')
                
                    t+=1 
                

            if line.split()[0]=="Section:":
           
                if t == 1:
                    grade=line.split()[1]

                    t+=1
                    
            if line.split()[0]=="Total":
                #print(line.split()[2])
                if line.split()[1]== "Membership:":
                    membership=int(line.split()[2])
                elif line.split()[1]== "Attendance:":
                    attendance=int(line.split()[2])
                    #print(y)
                    absent = membership-attendance 
                    #"# of students Absent: "+ 
                    percent= attendance/membership
                    percent=percent*100
                    if percent>int(login['threshold']):
                        abovex=1
                    else:
                        abovex=0
                    df = df.append({'Date':todayFormatted, 'Classroom':teacher,'Grade':grade,'Day Membership Possible':membership,'# of students Absent': absent, 'Day Attendance':attendance,'# of students Attending':math.ceil(percent), 'Above'+ login['threshold']: abovex},ignore_index = True)
                #print(str(absent)) , "% "+ "days above "+ login['threshold']: per
                t=0
    pdfFile.close()
    os.remove("ClassAttendanceAudit.pdf")
    os.chdir(os.path.realpath(os.path.join(os.path.dirname(__file__))))
   
    teacherList=df['Classroom'].values.tolist()
    gradeList=df['Grade'].values.tolist()
    above=df['Above'+ login['threshold']].values.tolist()
    if not forFile in book.sheetnames:
        book.create_sheet(forFile)
        #book.create_sheet('Graph By Teacher')
    try:
        book.remove_sheet(book['Sheet1'])   
    except:
        pass     
    with pd.ExcelWriter("classRoomAttendanceIncentive.xlsx", engine="openpyxl") as writer:
        writer.book = book
        writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))
        
        
            
        df.to_excel(writer,sheet_name = forFile,index=False, startrow=len(reader)+num)
    #if not 'Graph by Teacher' in book.sheetnames:
    try:       
        
        book.create_sheet('Graph By Teacher')
        book.save('classRoomAttendanceIncentive.xlsx')
        worksheet = book['Graph By Teacher'] 
        x=2
        for names in teacherList:
            cell = 'A'+ str(x)
            cellb = 'B'+str(x)
            cellc= 'C'+str(x)
            celle= 'E'+ str(x)
            worksheet[cell]=names
            worksheet[cellb]=0
            worksheet[cellc]=0
            worksheet[celle]=gradeList[x-2]
            x+=1
            #
        worksheet['B1']= 'Number of days above'+ login['threshold']
        worksheet['A1']= 'Class Room'
        worksheet['C1']= 'Total school Days'
        worksheet['D1']= '% days above'+ str(login['threshold'])
        worksheet['E1']= 'Grade'


        book.save('classRoomAttendanceIncentive.xlsx')
    except KeyError:
        print('Worksheet already exists')
        
        
    worksheet = book['Graph By Teacher']
    x=2
    for days in above:
            
        cellb = 'B'+str(x)
        cellc= 'C'+str(x)
        celld= 'D'+str(x)
        cellBval=worksheet[cellb].value
        cellCval=worksheet[cellc].value
        worksheet[cellb]=cellBval+above[x-2]
        worksheet[cellc]= cellCval+1
        worksheet[celld]=  (worksheet[cellb].value/ worksheet[cellc].value)*100

        x+=1
        book.save('classRoomAttendanceIncentive.xlsx')
    
    values= Reference(worksheet,
                      min_col=4,
                      max_col=4,
                      min_row=2,
                      max_row=34)
    cats = Reference(worksheet, min_col=1,max_col=1, min_row=2,max_row=34)
    chart = BarChart3D()
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    chart.title="Total Classroom Attendance %"
    chart.x_axis.title="Classroom"
    chart.x_axis.title="% Days Above"+login['threshold']
    worksheet.add_chart(chart,"G1")
    book.save('classRoomAttendanceIncentive.xlsx')
    book.close()
  


    
if __name__=="__main__":
    main()